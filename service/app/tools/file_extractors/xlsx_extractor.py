# service/app/tools/file_extractors/xlsx_extractor.py
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from .router import ExtractResult, sniff_mime
from .xlsx_image_map import extract_xlsx_anchored_images, load_anchored_image_bytes, a1

import zipfile


def extract_xlsx(
    *,
    filename: str,
    data: bytes,
    max_sheets: int = 4,
    max_rows: int = 140,
    max_cols: int = 30,
    max_chars: int = 60000,
    vision_caption_fn=None,
    max_images: int = 16,
) -> ExtractResult:
    """
    Deep XLSX extraction v1:
      - sheet-wise table extraction with cell coordinates (A1)
      - embedded images mapped to sheet + anchor cell using XML parsing
      - normalized JSON schema:
          extracted_json["sheets"] = [
            {
              "name": "Sheet1",
              "cells": [{"r":1,"c":1,"a1":"A1","v":"..."}, ...],
              "preview_text": "compact table preview",
              "images": [{"locator":"xlsx:FILE:sheet:Sheet1:img@B2","name":"image1.png","mime":"image/png","caption":"..."}]
            }, ...
          ]
    """
    meta: Dict[str, Any] = {
        "filename": filename,
        "max_sheets": int(max_sheets),
        "max_rows": int(max_rows),
        "max_cols": int(max_cols),
        "max_images": int(max_images),
    }

    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheetnames = wb.sheetnames[: int(max_sheets)]
        meta["sheets"] = sheetnames

        # --- Embedded image mapping (sheet + anchor cell) ---
        anchored = extract_xlsx_anchored_images(data, max_images=int(max_images))
        meta["embedded_images_anchored"] = len(anchored)
        img_bytes_map = load_anchored_image_bytes(data, anchored)

        # group images by sheet
        images_by_sheet: Dict[str, List[dict]] = {sn: [] for sn in sheetnames}
        for it in anchored:
            if it.sheet_name not in images_by_sheet:
                continue
            b = img_bytes_map.get(it.media_path, b"")
            mime = sniff_mime(filename=it.media_name, mime_type="", data=b) if b else "application/octet-stream"
            anchor = a1(it.col0, it.row0)  # top-left anchor
            loc = f"xlsx:{filename}:sheet:{it.sheet_name}:img@{anchor}"

            caption = ""
            if b and callable(vision_caption_fn):
                try:
                    caption = (vision_caption_fn(image_bytes=b, mime_type=mime, context=f"{filename}::{it.sheet_name}@{anchor}") or "").strip()
                except Exception as e:
                    meta.setdefault("vision_errors", []).append(str(e)[:200])
                    caption = ""

            images_by_sheet[it.sheet_name].append(
                {
                    "locator": loc,
                    "name": it.media_name,
                    "mime": mime,
                    "anchor_a1": anchor,
                    "caption": caption or "(No caption available.)",
                }
            )

        # --- Cell table extraction with coordinates ---
        sheets_out: List[Dict[str, Any]] = []
        chunks: List[str] = []

        for sn in sheetnames:
            ws = wb[sn]

            sheet_cells: List[Dict[str, Any]] = []
            preview_lines: List[str] = []
            preview_lines.append(f"--- SHEET: {sn} (xlsx:{filename}:sheet:{sn}) ---")

            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if r_idx > int(max_rows):
                    break
                row_vals: List[str] = []
                for c_idx, v in enumerate(row, start=1):
                    if c_idx > int(max_cols):
                        break
                    val = "" if v is None else str(v).strip()
                    if val == "":
                        continue

                    sheet_cells.append(
                        {"r": r_idx, "c": c_idx, "a1": a1(c_idx - 1, r_idx - 1), "v": val}
                    )
                    # preview row (compact)
                    row_vals.append(f"{a1(c_idx - 1, r_idx - 1)}={val}")

                if row_vals:
                    preview_lines.append(" | ".join(row_vals))

                if sum(len(x) for x in preview_lines) > max_chars:
                    break

            # include image captions in preview
            imgs = images_by_sheet.get(sn, []) or []
            if imgs:
                preview_lines.append("--- EMBEDDED IMAGES ---")
                for im in imgs[:12]:
                    preview_lines.append(f"{im.get('locator')}: {im.get('caption')}")

            preview_text = "\n".join(preview_lines).strip()
            chunks.append(preview_text)

            sheets_out.append(
                {
                    "name": sn,
                    "locator": f"xlsx:{filename}:sheet:{sn}",
                    "cells": sheet_cells[: 12000],  # safety bound
                    "preview_text": preview_text[:8000],
                    "images": imgs,
                }
            )

            if sum(len(x) for x in chunks) > max_chars:
                break

        wb.close()

        text = "\n\n".join(chunks).strip()
        if len(text) > max_chars:
            text = text[:max_chars] + "\n\n[TRUNCATED]"
        if not text:
            text = "(Empty XLSX.)"

        return ExtractResult(
            doc_type="xlsx",
            extracted_text=text,
            extracted_json={"sheets": sheets_out},
            meta=meta,
        )

    except Exception as e:
        return ExtractResult(
            doc_type="xlsx",
            extracted_text="(XLSX extraction failed.)",
            extracted_json={},
            meta={**meta, "error": str(e)[:300]},
        )